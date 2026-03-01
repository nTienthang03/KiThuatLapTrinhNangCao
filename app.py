from __future__ import annotations

from datetime import datetime
from io import BytesIO

from flask import Flask, redirect, render_template, request, url_for
from flask import send_file

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from data import get_connection
from optimizer import (
    ensure_missing_invigilators,
    fetch_room_details,
    fetch_schedule_summary,
    fetch_student_schedules,
    optimize_schedule,
)
from seed import seed_sample_data

app = Flask(__name__)


def _mode_label(mode: str) -> str:
    return {
        "spread": "phan-bo-deu",
        "compact": "toi-uu-phong-thoi-gian",
        "balanced": "can-bang",
    }.get(mode, "can-bang")


def _schedule_to_xlsx_bytes(schedule: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "LichThi"

    headers = [
        ("MaLich", "Mã lịch"),
        ("MaMon", "Mã môn"),
        ("TenMon", "Tên môn"),
        ("NgayThi", "Ngày thi"),
        ("ThuTuTrongNgay", "Ca"),
        ("GioBatDau", "Giờ bắt đầu"),
        ("GioKetThuc", "Giờ kết thúc"),
        ("MaPhong", "Mã phòng"),
        ("TenPhong", "Tên phòng"),
        ("SoSinhVien", "Số SV"),
        ("GhiChu", "Ghi chú"),
    ]

    ws.append([h[1] for h in headers])
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in schedule:
        ws.append([row.get(k, "") for k, _ in headers])

    # Basic widths
    widths = [8, 10, 28, 12, 6, 12, 12, 10, 18, 8, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/")
def index():
    min_gap_days = 0
    with get_connection() as conn:
        ensure_missing_invigilators(conn)
        schedule = fetch_schedule_summary(conn)
        room_details = fetch_room_details(conn)
        student_schedules = fetch_student_schedules(conn)
        conn.commit()
    return render_template(
        "index.html",
        schedule=schedule,
        room_details=room_details,
        student_schedules=student_schedules,
        message=None,
        is_error=False,
        min_gap_days=min_gap_days,
        mode="balanced",
        stats=None,
    )


@app.post("/optimize")
def optimize():
    mode = request.form.get("mode", "balanced")
    if mode not in ("spread", "compact", "balanced"):
        mode = "balanced"
    try:
        min_gap_days = int(request.form.get("min_gap_days", "0"))
    except ValueError:
        min_gap_days = 0

    message = None
    is_error = False
    stats = None

    try:
        with get_connection() as conn:
            result = optimize_schedule(
                conn, mode=mode, min_gap_days=min_gap_days, clear_existing=True,
            )
            conn.commit()

            ensure_missing_invigilators(conn)
            conn.commit()

            stats = {
                "created_lich": result.created_lich,
                "created_assignments": result.created_assignments,
                "total_exam_days": result.total_exam_days,
                "total_slots_used": result.total_slots_used,
            }

            if result.unscheduled_mon:
                message = (
                    f"Đã tạo {result.created_lich} lịch, gán {result.created_assignments} lượt SV. "
                    f"Chưa xếp được môn: {', '.join(result.unscheduled_mon)}"
                )
                is_error = True
            else:
                message = (
                    f"Đã tạo {result.created_lich} lịch, gán {result.created_assignments} lượt SV "
                    f"trong {result.total_exam_days} ngày, {result.total_slots_used} ca."
                )

            schedule = fetch_schedule_summary(conn)
            room_details = fetch_room_details(conn)
            student_schedules = fetch_student_schedules(conn)

    except Exception as exc:  # noqa: BLE001
        message = f"Lỗi khi tối ưu: {exc}"
        is_error = True
        with get_connection() as conn:
            ensure_missing_invigilators(conn)
            schedule = fetch_schedule_summary(conn)
            room_details = fetch_room_details(conn)
            student_schedules = fetch_student_schedules(conn)
            conn.commit()

    return render_template(
        "index.html",
        schedule=schedule,
        room_details=room_details,
        student_schedules=student_schedules,
        message=message,
        is_error=is_error,
        min_gap_days=min_gap_days,
        mode=mode,
        stats=stats,
    )


@app.get("/optimize")
def optimize_get():
    return redirect(url_for("index"))


@app.post("/seed")
def seed():
    mode = request.form.get("mode", "balanced")
    if mode not in ("spread", "compact", "balanced"):
        mode = "balanced"
    try:
        min_gap_days = int(request.form.get("min_gap_days", "0"))
    except ValueError:
        min_gap_days = 0

    reset = request.form.get("reset_seed", "0") == "1"

    message = None
    is_error = False
    stats = None

    try:
        with get_connection() as conn:
            r = seed_sample_data(conn, reset=reset)
            conn.commit()
            ensure_missing_invigilators(conn)
            schedule = fetch_schedule_summary(conn)
            room_details = fetch_room_details(conn)
            student_schedules = fetch_student_schedules(conn)
            conn.commit()

        message = (
            "Đã tạo dữ liệu mẫu" + (" (reset)" if reset else "") + ": "
            f"+{r.inserted_students} SV, +{r.inserted_courses} môn, +{r.inserted_rooms} phòng, "
            f"+{r.inserted_invigilators} giám thị, "
            f"+{r.inserted_slots} ca, +{r.inserted_availability} khả dụng, +{r.inserted_registrations} đăng ký."
        )
    except Exception as exc:  # noqa: BLE001
        message = f"Lỗi khi tạo dữ liệu mẫu: {exc}"
        is_error = True
        with get_connection() as conn:
            schedule = fetch_schedule_summary(conn)
            room_details = fetch_room_details(conn)
            student_schedules = fetch_student_schedules(conn)
            conn.commit()

    return render_template(
        "index.html",
        schedule=schedule,
        room_details=room_details,
        student_schedules=student_schedules,
        message=message,
        is_error=is_error,
        min_gap_days=min_gap_days,
        mode=mode,
        stats=stats,
    )


@app.get("/seed")
def seed_get():
    return redirect(url_for("index"))


@app.post("/export")
def export_excel():
    mode = request.args.get("mode", "balanced")
    if mode not in ("spread", "compact", "balanced"):
        mode = "balanced"

    try:
        min_gap_days = int(request.form.get("min_gap_days", "0"))
    except ValueError:
        min_gap_days = 0

    # Generate inside a transaction and rollback so exporting does not change stored schedule.
    with get_connection() as conn:
        try:
            optimize_schedule(conn, mode=mode, min_gap_days=min_gap_days, clear_existing=True)
            ensure_missing_invigilators(conn)
            schedule = fetch_schedule_summary(conn)
        finally:
            conn.rollback()

    xlsx = _schedule_to_xlsx_bytes(schedule)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lich_thi_{_mode_label(mode)}_{timestamp}.xlsx"

    return send_file(
        xlsx,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/export")
def export_get():
    return redirect(url_for("index"))


def _sheet_apply_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")


@app.get("/export_view")
def export_view_excel():
    view = request.args.get("view", "schedule")
    if view not in ("schedule", "rooms", "students"):
        view = "schedule"

    with get_connection() as conn:
        ensure_missing_invigilators(conn)
        schedule = fetch_schedule_summary(conn)
        room_details = fetch_room_details(conn)
        student_schedules = fetch_student_schedules(conn)
        conn.commit()

    wb = Workbook()
    # remove default empty sheet
    default_ws = wb.active
    wb.remove(default_ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if view == "schedule":
        ws = wb.create_sheet("LichThi")
        headers = [
            "MaLich",
            "MaMon",
            "TenMon",
            "NgayThi",
            "Ca",
            "GioBatDau",
            "GioKetThuc",
            "MaPhong",
            "TenPhong",
            "SoSinhVien",
            "SucChua",
            "PhanTramLap",
            "SoGiamThi",
            "GhiChu",
        ]
        _sheet_apply_header(ws, headers)
        for row in schedule:
            det = room_details.get(row["MaLich"], {})
            so_gt = len(det.get("giamthi", [])) if det else 0
            ws.append(
                [
                    row.get("MaLich"),
                    row.get("MaMon"),
                    row.get("TenMon"),
                    row.get("NgayThi"),
                    row.get("ThuTuTrongNgay"),
                    row.get("GioBatDau"),
                    row.get("GioKetThuc"),
                    row.get("MaPhong"),
                    row.get("TenPhong"),
                    row.get("SoSinhVien"),
                    row.get("SucChua"),
                    row.get("PhanTramLap"),
                    so_gt,
                    row.get("GhiChu"),
                ]
            )
        xlsx = BytesIO()
        wb.save(xlsx)
        xlsx.seek(0)
        filename = f"lich_thi_{timestamp}.xlsx"
        return send_file(
            xlsx,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if view == "rooms":
        ws_gt = wb.create_sheet("GiamThi")
        _sheet_apply_header(ws_gt, [
            "MaLich",
            "MaMon",
            "TenMon",
            "NgayThi",
            "Ca",
            "MaPhong",
            "TenPhong",
            "ThuTu",
            "MaGiamThi",
            "HoTen",
            "DonVi",
        ])

        ws_sv = wb.create_sheet("ThiSinh")
        _sheet_apply_header(ws_sv, [
            "MaLich",
            "MaMon",
            "TenMon",
            "NgayThi",
            "Ca",
            "MaPhong",
            "TenPhong",
            "MaSinhVien",
            "HoTen",
            "Lop",
        ])

        by_lich = {r["MaLich"]: r for r in schedule}
        for ma_lich, det in room_details.items():
            s = by_lich.get(ma_lich)
            if not s:
                continue
            for gt in det.get("giamthi", []):
                ws_gt.append([
                    ma_lich,
                    s.get("MaMon"),
                    s.get("TenMon"),
                    s.get("NgayThi"),
                    s.get("ThuTuTrongNgay"),
                    s.get("MaPhong"),
                    s.get("TenPhong"),
                    gt.get("ThuTu"),
                    gt.get("MaGiamThi"),
                    gt.get("HoTen"),
                    gt.get("DonVi"),
                ])
            for sv in det.get("sinhvien", []):
                ws_sv.append([
                    ma_lich,
                    s.get("MaMon"),
                    s.get("TenMon"),
                    s.get("NgayThi"),
                    s.get("ThuTuTrongNgay"),
                    s.get("MaPhong"),
                    s.get("TenPhong"),
                    sv.get("MaSinhVien"),
                    sv.get("HoTen"),
                    sv.get("Lop"),
                ])

        xlsx = BytesIO()
        wb.save(xlsx)
        xlsx.seek(0)
        filename = f"chi_tiet_phong_{timestamp}.xlsx"
        return send_file(
            xlsx,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # view == students
    ws = wb.create_sheet("LichTheoSV")
    _sheet_apply_header(ws, [
        "MaSinhVien",
        "HoTen",
        "Lop",
        "MaLich",
        "MaMon",
        "TenMon",
        "NgayThi",
        "Ca",
        "GioBatDau",
        "GioKetThuc",
        "MaPhong",
        "TenPhong",
    ])
    for sv in student_schedules:
        for ex in sv.get("exams", []):
            ws.append([
                sv.get("MaSinhVien"),
                sv.get("HoTen"),
                sv.get("Lop"),
                ex.get("MaLich"),
                ex.get("MaMon"),
                ex.get("TenMon"),
                ex.get("NgayThi"),
                ex.get("ThuTuTrongNgay"),
                ex.get("GioBatDau"),
                ex.get("GioKetThuc"),
                ex.get("MaPhong"),
                ex.get("TenPhong"),
            ])

    xlsx = BytesIO()
    wb.save(xlsx)
    xlsx.seek(0)
    filename = f"lich_theo_sinh_vien_{timestamp}.xlsx"
    return send_file(
        xlsx,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    # use_reloader=False để tránh parent process thoát làm bạn tưởng app đã dừng.
    app.run(host="127.0.0.1", port=5000, debug=True, use_reloader=False)
